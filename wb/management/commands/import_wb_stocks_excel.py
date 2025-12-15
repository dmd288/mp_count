import json
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from wb.models import ImportFile, ImportRow, WBProduct, WBBarcode, WBStockSnapshot


class Command(BaseCommand):
    help = 'Import Wildberries stock data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to the Excel file to import'
        )

    def handle(self, *args, **options):
        file_path = Path(options['file'])
        
        if not file_path.exists():
            raise CommandError(f'File not found: {file_path}')
        
        if not file_path.suffix.lower() in ['.xlsx', '.xls']:
            raise CommandError(f'File must be an Excel file (.xlsx or .xls): {file_path}')
        
        self.stdout.write(f'Loading Excel file: {file_path}')
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Find header row
            header_row = None
            header_mapping = {}
            
            # Expected Russian column headers
            expected_headers = {
                'Бренд': 'brand',
                'Предмет': 'subject',
                'Артикул продавца': 'vendor_code',
                'Артикул WB': 'nm_id',
                'Объем, л': 'volume',
                'Баркод': 'barcode',
                'Размер вещи': 'tech_size',
                'В пути до получателей': 'in_way_to_client',
                'В пути возвраты на склад WB': 'in_way_from_client',
                'Всего находится на складах': 'quantity_full',
            }
            
            # Try to find header row (first 5 rows)
            for row_idx in range(1, min(6, worksheet.max_row + 1)):
                row = worksheet[row_idx]
                found_headers = {}
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip()
                        if cell_value in expected_headers:
                            found_headers[expected_headers[cell_value]] = cell.column - 1  # 0-indexed
                
                if len(found_headers) >= 3:  # At least 3 required columns found
                    header_row = row_idx
                    header_mapping = found_headers
                    break
            
            if header_row is None:
                error_msg = (
                    f"Could not find required column headers. "
                    f"Expected at least: Артикул продавца, Артикул WB, Баркод"
                )
                self._create_error_import_file(file_path, error_msg)
                raise CommandError(error_msg)
            
            # Check for required columns
            required_columns = ['vendor_code', 'nm_id', 'barcode']
            missing_columns = [col for col in required_columns if col not in header_mapping]
            
            if missing_columns:
                russian_names = {
                    'vendor_code': 'Артикул продавца',
                    'nm_id': 'Артикул WB',
                    'barcode': 'Баркод',
                }
                missing_russian = [russian_names[col] for col in missing_columns]
                error_msg = f"Missing required columns: {', '.join(missing_russian)}"
                self._create_error_import_file(file_path, error_msg)
                raise CommandError(error_msg)
            
            # Create ImportFile record
            import_file = ImportFile(
                source='WB_STOCKS_EXCEL',
                original_filename=file_path.name,
                uploaded_at=timezone.now(),
                status='PARSED'
            )
            # Save file to FileField
            with open(file_path, 'rb') as f:
                import_file.file.save(file_path.name, File(f), save=False)
            import_file.save()
            
            self.stdout.write(f'Created ImportFile: {import_file.id}')
            
            # Process data rows
            row_count = 0
            error_count = 0
            
            with transaction.atomic():
                for row_idx in range(header_row + 1, worksheet.max_row + 1):
                    row = worksheet[row_idx]
                    row_data = {}
                    row_errors = []
                    
                    # Extract data from row
                    for field_name, col_idx in header_mapping.items():
                        try:
                            cell = row[col_idx]
                            value = cell.value
                            
                            # Convert to appropriate type
                            if field_name in ['nm_id', 'quantity', 'in_way_to_client', 
                                            'in_way_from_client', 'quantity_full']:
                                if value is None:
                                    value = 0
                                else:
                                    try:
                                        value = int(float(str(value)))
                                    except (ValueError, TypeError):
                                        value = 0
                            elif field_name == 'volume':
                                if value is None:
                                    value = None
                                else:
                                    try:
                                        value = float(str(value))
                                    except (ValueError, TypeError):
                                        value = None
                            else:
                                value = str(value).strip() if value else ''
                            
                            row_data[field_name] = value
                        except (IndexError, AttributeError) as e:
                            row_errors.append(f"Error reading column {field_name}: {str(e)}")
                    
                    # Validate required fields
                    if not row_data.get('nm_id') or row_data.get('nm_id') == 0:
                        row_errors.append("Missing or invalid nm_id (Артикул WB)")
                    if not row_data.get('vendor_code'):
                        row_errors.append("Missing vendor_code (Артикул продавца)")
                    if not row_data.get('barcode'):
                        row_errors.append("Missing barcode (Баркод)")
                    
                    # Skip empty rows
                    if not any(row_data.values()):
                        continue
                    
                    # Create ImportRow
                    import_row = ImportRow.objects.create(
                        import_file=import_file,
                        row_number=row_idx,
                        raw_json=row_data,
                        errors=row_errors if row_errors else None
                    )
                    
                    if row_errors:
                        error_count += 1
                        continue
                    
                    # Upsert WBProduct
                    nm_id = int(row_data['nm_id'])
                    vendor_code = str(row_data['vendor_code']).strip()
                    
                    wb_product, created = WBProduct.objects.update_or_create(
                        nm_id=nm_id,
                        defaults={
                            'vendor_code': vendor_code,
                            'brand': row_data.get('brand', ''),
                            'subject': row_data.get('subject', ''),
                        }
                    )
                    
                    # Create/update WBBarcode
                    barcode = str(row_data['barcode']).strip()
                    tech_size = row_data.get('tech_size', '').strip()
                    
                    wb_barcode, _ = WBBarcode.objects.update_or_create(
                        barcode=barcode,
                        defaults={
                            'wb_product': wb_product,
                            'tech_size': tech_size,
                        }
                    )
                    
                    # Create WBStockSnapshot
                    warehouse_name = row_data.get('warehouse_name', '').strip()
                    in_way_to_client = row_data.get('in_way_to_client', 0) or 0
                    in_way_from_client = row_data.get('in_way_from_client', 0) or 0
                    quantity_full = row_data.get('quantity_full', 0) or 0
                    # quantity is not in Excel, so we set it equal to quantity_full
                    quantity = quantity_full
                    
                    WBStockSnapshot.objects.create(
                        import_file=import_file,
                        warehouse_name=warehouse_name,
                        nm_id=nm_id,
                        vendor_code=vendor_code,
                        barcode=barcode,
                        tech_size=tech_size,
                        quantity=quantity,
                        in_way_to_client=in_way_to_client,
                        in_way_from_client=in_way_from_client,
                        quantity_full=quantity_full,
                    )
                    
                    row_count += 1
                    
                    if row_count % 100 == 0:
                        self.stdout.write(f'Processed {row_count} rows...')
                
                # Update import file status
                if error_count > 0:
                    import_file.status = 'APPLIED'
                    import_file.error_log = f"Processed {row_count} rows successfully, {error_count} rows with errors"
                else:
                    import_file.status = 'APPLIED'
                    import_file.error_log = f"Successfully processed {row_count} rows"
                
                import_file.save()
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'Successfully imported {row_count} rows from {file_path.name}. '
                    f'Errors: {error_count}'
                )
            )
            
        except Exception as e:
            error_msg = f"Error processing file: {str(e)}"
            self.stdout.write(self.style.ERROR(error_msg))
            self._create_error_import_file(file_path, error_msg)
            raise CommandError(error_msg)
    
    def _create_error_import_file(self, file_path, error_msg):
        """Create an ImportFile record with ERROR status"""
        try:
            ImportFile.objects.create(
                source='WB_STOCKS_EXCEL',
                original_filename=file_path.name,
                file=str(file_path),
                status='ERROR',
                error_log=error_msg,
            )
        except Exception:
            pass  # If we can't create the record, just continue

